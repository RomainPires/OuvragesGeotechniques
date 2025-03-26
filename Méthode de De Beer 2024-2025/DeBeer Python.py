import numpy as np
import pandas as pd
import matplotlib.pyplot as plt
from openpyxl import load_workbook

def Calcul_Beta(h_d, Phip):
    Beta_min = 0
    Beta_max = np.pi

    h_d_1 = Calcul_h_d((Beta_min + Beta_max) / 2, Phip)

    while abs(h_d - h_d_1) > 0.001:
        if h_d < h_d_1:
            Beta_max = (Beta_min + Beta_max) / 2
        else:
            Beta_min = (Beta_min + Beta_max) / 2

        h_d_1 = Calcul_h_d((Beta_min + Beta_max) / 2, Phip)

        if (Beta_max > np.pi / 2 and Beta_min > np.pi / 2):
            Beta_max = np.pi / 2
            Beta_min = np.pi / 2
            h_d_1 = h_d

    return (Beta_min + Beta_max) / 2

def Calcul_h_d(beta, Phip):
    return (np.tan(np.pi / 4 + Phip / 2) * np.exp(np.pi / 2 * np.tan(Phip)) * np.sin(beta) * np.exp(beta * np.tan(Phip))) / (1 + np.sin(2 * Phip))

def Calcul_Phip(Vpbd, Phi):
    Phip_min = 0
    Phip_max = 60 * np.pi / 180

    Zeroabs = 1.3 * 2 * np.pi - 0.5  # Valeur minimum de Vbd pour Phi=0
    if Vpbd <= Zeroabs:
        return 0

    Vpbd_1 = Calcul_Vpbd((Phip_min + Phip_max) / 2, Phi)

    while abs(Vpbd - Vpbd_1) > 0.001:
        if Vpbd < Vpbd_1:
            Phip_max = (Phip_min + Phip_max) / 2
        else:
            Phip_min = (Phip_min + Phip_max) / 2

        Vpbd_1 = Calcul_Vpbd((Phip_min + Phip_max) / 2, Phi)

    if (Phip_min + Phip_max) / 2 < Phi:
        return (Phip_min + Phip_max) / 2
    else:
        return Calcul_Phi(Vpbd)

def Calcul_Vpbd(Phip, Phi):
    if Phip <= 0:
        return 1.3 * (2 * np.pi - 0.5)  # valeur limite
    else:
        return 1.3 * ((np.exp(2 * np.pi * np.tan(Phip)) * (np.tan(np.pi / 4 + Phip / 2)) ** 2 - 1) * np.tan(Phi) / np.tan(Phip) + 1)  # inversion Phi - Phip !!

def Calcul_Phi(Vbd):
    if Vbd < 1.3:
        return -999
    else:
        Phi_min = 0
        Phi_max = 60 * np.pi / 180

        Vbd_1 = Calcul_Vbd((Phi_min + Phi_max) / 2)

        while abs(Vbd - Vbd_1) > 0.001:
            if Vbd < Vbd_1:
                Phi_max = (Phi_min + Phi_max) / 2
            else:
                Phi_min = (Phi_min + Phi_max) / 2

            Vbd_1 = Calcul_Vbd((Phi_min + Phi_max) / 2)

        return (Phi_min + Phi_max) / 2

def Calcul_Vbd(Phi):
    return 1.3 * np.exp(2 * np.pi * np.tan(Phi)) * (np.tan(np.pi / 4 + Phi / 2)) ** 2


########### DE BEER CODE 
# Initialisation
Color = ['#0072BD', '#D95319', '#77AC30', '#A2142F', '#0072BD', '#D95319', '#EDB120', '#77AC30', '#A2142F', '#0072BD', '#D95319', '#EDB120', '#77AC30', '#A2142F', '#0072BD', '#D95319', '#EDB120', '#77AC30', '#A2142F']
Color2 = np.array([[0, 0.4470, 0.7410], [0.8500, 0.3250, 0.0980], [0.9290, 0.6940, 0.1250], [0.4660, 0.6740, 0.1880], [0.6350, 0.0780, 0.1840]])

Line = ['-', '--', '-.', ':', '-', '--', '-.', '-', '--', '-.', '-', '--', '-.', '-', '--', '-.', '-', '--']
Marker = ['o', '^', 's', '*', '<', '>', 'o', '^', 's', '*', '<', '>', 'o', '^', 's', '*', '<', '>']

plt.rcParams['axes.labelsize'] = 30
plt.rcParams['axes.titlesize'] = 30
plt.rcParams['axes.linewidth'] = 2.5
plt.rcParams['lines.linewidth'] = 3

# Données de base
Gamma_w = 1000  # kg/m³
Gamma_sat = 2000  # kg/m³
Gamma_h = 1600  # kg/m³

d = 35.7  # mm

Tiges = 0.0140  # MPa
Cone = 0.0254  # MPa

Epaisseur = 1.5  # m

# Données utilisateur
D = 600  # mm
Zw = 3.3  # m

# Importation de Z et qc
ExcelDeBeer = pd.read_excel("/workspaces/OuvragesGeotechniques/Méthode de De Beer 2024-2025/MethodeDeBeer.xlsx", sheet_name="Calculs", usecols="A:B")

# Iterate through the transposed DataFrame to extract each column
columns = {}
for column_name, column_data in ExcelDeBeer.items():
    columns[column_name] = column_data.tolist()

Z =  columns['Z'][1:] #m
qc = columns['qc'][1:] #MPa

MeanZdiff = (Z[-1]-Z[0])/len(Z)

MaxZReal = Z[-1]
RealZlen = len(Z)

for i in range(10) : # Prolongation
    Z.append(Z[-1]+MeanZdiff)
    qc.append(qc[-1])

# Calcul
Nmax = 0
ZMax = 0
pb = np.zeros(len(Z))
Phi = np.zeros(len(Z))
Phi_p = np.zeros(len(Z))
h_d = np.zeros(len(Z))
Beta_c = np.zeros(len(Z))
h_D = np.zeros(len(Z))
Beta_p = np.zeros(len(Z))
dg1 = np.zeros(len(Z))
hcrit = np.zeros(len(Z))
hpcrit = np.zeros(len(Z))
dgj1 = np.zeros(len(Z))
A = np.zeros(len(Z))
dg = np.zeros(len(Z))
dgq1 = np.zeros(len(Z))
dg_save = np.zeros(len(Z))

for i in range(len(Z)):
    if Z[i] == 0:
        break
    else:
        if Z[i] <= Zw:
            Gamma = Gamma_h * 9.81
        else:
            Gamma = (Gamma_sat - Gamma_w) * 9.81

        Ckd = qc[i]

        if i == 0:
            pb[i] = (Z[i] - ZMax) * Gamma / 1000  # kPa
        else:
            pb[i] = pb[i - 1] + (Z[i] - ZMax) * Gamma / 1000  # kPa

        Phi[i] = 30 * np.pi / 180
        Phi_p[i] = Calcul_Phip(Ckd / 0.001 / pb[i], Phi[i])

        h_d[i] = Z[i] / d * 1000

        Beta_c[i] = Calcul_Beta(h_d[i], Phi_p[i])

        h_D[i] = Z[i] / D * 1000

        Beta_p[i] = Calcul_Beta(h_D[i], Phi_p[i])

        dg1[i] = Ckd / np.exp(2 * (Beta_c[i] - Beta_p[i]) * np.tan(Phi_p[i]))

        ZMax = Z[i]
        Nmax += 1

Phi = np.rad2deg(Phi)
Phi_p = np.rad2deg(Phi_p)

for i in range(1, Nmax + 1):
    if i > 2:
        Save = [Phi_p[i - 3], Phi_p[i - 2], Phi_p[i - 1]]
        
        if any(s > 37.5 for s in Save):
            Ihcrit = 1
        elif any(s > 32.5 for s in Save):
            Ihcrit = 2
        else:
            Ihcrit = 1
    else:
        Ihcrit = 1
    
    if Ihcrit == 1:
        hcrit[i - 1] = MeanZdiff#0.2
        hpcrit[i - 1] = hcrit[i - 1] * D / d
        
        if i != 1:
            pbm1 = pb[i - 2]
        else:
            pbm1 = 0
        
        if Z[i - 1] <= Zw:
            Gamma = Gamma_h
        else:
            Gamma = Gamma_sat - Gamma_w
        
        A[i - 1] = (pbm1 + 0.5 * Gamma * 9.81 / 1000 * hpcrit[i - 1]) / (pbm1 + 0.5 * Gamma * 9.81 / 1000 * hcrit[i - 1])

        if i != 1:
            dgjm1 = dgj1[i - 2]
        else:
            dgjm1 = 0
        
        dgj1[i - 1] = dgjm1 + d / D * (A[i - 1] * dg1[i - 1] - dgjm1)
        
        if dgj1[i - 1] > dg1[i - 1]:
            dgj1[i - 1] = dg1[i - 1]
    
    elif Ihcrit == 2:
        hcrit_1 = MeanZdiff#0.2
        hpcrit_1 = hcrit_1 * D / d
        if i != 1:
            pbm1_1 = pb[i - 2]
            if Z[i - 1] <= Zw:
                Gamma = Gamma_h
            else:
                Gamma = Gamma_sat - Gamma_w
            A_1 = (pbm1_1 + 0.5 * Gamma * 9.81 / 1000 * hpcrit_1) / (pbm1_1 + 0.5 * Gamma * 9.81 / 1000 * hcrit_1)
        else:
            A_1 = 0
        
        if i != 1:
            dgjm1_1 = dgj1[i - 2]
            dg1_1 = dg[i - 1]
            dgj1_1 = dgjm1_1 + d / D * (A_1 * dg1_1 - dgjm1_1)
        else:
            dgj1_1 = dg[i - 1]
        
        hcrit_2 = 2*MeanZdiff #0.4
        hpcrit_2 = hcrit_2 * D / d
        
        if i != 1:
            pbm1_2 = pb[i - 2]
            if Z[i - 1] <= Zw:
                Gamma = Gamma_h
            else:
                Gamma = Gamma_sat - Gamma_w
            A_2 = (pbm1_2 + 0.5 * Gamma * 9.81 / 1000 * hpcrit_2) / (pbm1_2 + 0.5 * Gamma * 9.81 / 1000 * hcrit_2)
        else:
            A_2 = 0
        
        if i != 1:
            dgjm1_2 = dgj1[i - 2]
            dg1_2 = dg[i - 1]
            dgj1_2 = dgjm1_2 + 0.5 * d / D * (A_2 * dg1_2 - dgjm1_2)
        else:
            dgj1_2 = dg1_2
        
        if dgj1_1 < dgj1_2:
            hcrit[i - 1] = hcrit_1
            hpcrit[i - 1] = hpcrit_1
            A[i - 1] = A_1
            if dgj1_1 > dg1_1:
                dgj1[i - 1] = dg1_1
            else:
                dgj1[i - 1] = dgj1_1
        else:
            hcrit[i - 1] = hcrit_2
            hpcrit[i - 1] = hpcrit_2
            A[i - 1] = A_2
            if dgj1_2 > dg1_1:
                dgj1[i - 1] = dg1_1
            else:
                dgj1[i - 1] = dgj1_2

for i in range(Nmax, 0, -1):
    if i == Nmax:
        dgq1[i - 1] = dgj1[i - 1]
    else:
        dgq1[i - 1] = dgq + (dgj1[i - 1] - dgq) / (D / d)

        if dgj1[i - 1] < dgq:
            dgq1[i - 1] = dgj1[i - 1]
    dgq = dgq1[i - 1]

for i in range(Nmax):
    DZ = round(D / (MeanZdiff*1000) + 0.5)
    Somme = 0
    
    if Z[i] + (DZ * MeanZdiff) < ZMax: # REPLACE 0.2 with MeanZdiff ???
        for j in range(i, min(i + DZ + 1, Nmax + 1)):
            Somme += dgq1[j - 1]
    
    dg = Somme / (DZ + 1)
    if dg > dg1[i]:
        dg = dg1[i]
    if dg > 0:
        dg_save[i] = dg

#for i in range(len(dg_save)) :
#    print(dg_save[i])

Z = Z[0:RealZlen]
dg_save = dg_save[0:RealZlen]
qc = qc[0:RealZlen]

# Sauvetage dans l'excel
wb = load_workbook("/workspaces/OuvragesGeotechniques/Méthode de De Beer 2024-2025/MethodeDeBeer.xlsx")
ws = wb["Calculs"]

for i in range(len(Z)):
    ws[f"C{i+3}"] = Z[i]
    ws[f"D{i+3}"] = dg_save[i]

# Save the changes to the Excel file
wb.save("MethodeDeBeer.xlsx")

# Plot
plt.figure()
plt.plot(qc, Z, '-', linewidth=2)
plt.plot(dg_save, Z[:len(dg_save)], '-', linewidth=2, alpha=0.8)
plt.xlabel('q_c [MPa]', fontsize=12)
plt.ylabel('Profondeur [m]', fontsize=12)
plt.gca().invert_yaxis()
plt.grid(True)
plt.title("Portance limite en base d'un pieu de diamètre D=800 mm", fontsize=14)
plt.legend(['q_c CPT', 'q_c lissé De Beer'], loc='upper right', fontsize=10)
plt.show()
